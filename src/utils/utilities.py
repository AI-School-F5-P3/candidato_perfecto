"""Funciones de utilidad para el sistema de análisis de RRHH"""
import logging
from typing import List, Dict, Any, Union
import pandas as pd
import json
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

def setup_logging(log_file: str = "app.log") -> None:
    """Configura el sistema de registro para la aplicación"""
    # Configura el registro tanto en archivo como en consola
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s: %(message)s",
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )

def format_list_preview(items: List[str], max_items: int = 5) -> str:
    """Formatea una lista para vista previa, mostrando solo los primeros elementos"""
    preview = items[:max_items]
    # Agrega "..." si hay más elementos que el máximo mostrado
    return ', '.join(preview) + ('...' if len(items) > max_items else '')

def create_score_row(candidate_data: Dict[str, Any], score_data: Dict[str, Any]) -> Dict[str, Any]:
    """Crea una fila de datos para el DataFrame de resultados"""
    import json
    # Añadir campos esenciales
    candidate_data.setdefault('job_skills', [])
    candidate_data.setdefault('job_experience', 0)
    candidate_data.setdefault('preferencias_score', 0.0)

    # Convertir experiencia a strings para visualización
    exp_preview = []
    for exp in candidate_data.get('experiencia', []):
        if isinstance(exp, dict):
            exp_str = f"{exp.get('puesto', '')} ({exp.get('duracion', '')})"
            exp_preview.append(exp_str)
        else:
            exp_preview.append(str(exp))

    # Convert nested dicts to strings for safe DataFrame creation.
    if "raw_data" in candidate_data and isinstance(candidate_data["raw_data"], dict):
        candidate_data["raw_data"] = json.dumps(candidate_data["raw_data"])
    # Ensure candidate list fields contain only strings.
    for key in ['habilidades', 'experiencia', 'formacion']:
        if key in candidate_data and isinstance(candidate_data[key], list):
            candidate_data[key] = [json.dumps(item) if isinstance(item, dict) else str(item) for item in candidate_data[key]]
    # Construye un diccionario con los datos formateados para visualización
    return {
        'Nombre Candidato': candidate_data['nombre_candidato'],
        'Estado': 'Descalificado' if score_data['disqualified'] else 'Calificado',
        'Score Final': f"{score_data['final_score']:.1%}",
        'Score Habilidades': f"{score_data['component_scores']['habilidades']:.1%}",
        'Score Experiencia': f"{score_data['component_scores']['experiencia']:.1%}",
        'Score Formación': f"{score_data['component_scores']['formacion']:.1%}",
        'Score Preferencias': f"{score_data['component_scores']['preferencias_reclutador']:.1%}",
        'Habilidades': format_list_preview(candidate_data['habilidades']),
        'Experiencia': format_list_preview(exp_preview, 3),
        'Formación': format_list_preview(candidate_data['formacion'], 2),
        'Razones Descalificación': ', '.join(score_data.get('disqualification_reasons', [])) or 'N/A',
        'raw_data': json.dumps(candidate_data)  # Asegurar serialización
    }

def sort_ranking_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Ordena el DataFrame de ranking por puntuación y estado de descalificación"""
    # Convierte las puntuaciones de porcentaje a números para ordenamiento
    df['Sort Score'] = df['Score Final'].str.rstrip('%').astype('float')
    
    # Ordena primero por estado (Calificado antes que Descalificado) y luego por puntuación
    df = df.sort_values(
        by=['Estado', 'Sort Score'], 
        ascending=[True, False],
        key=lambda x: x if x.name != 'Estado' else pd.Categorical(x, ['Calificado', 'Descalificado'])
    )
    return df.drop('Sort Score', axis=1)  # Elimina la columna temporal de ordenamiento

def export_rankings_to_excel(rankings_dfs, sheet_names, buffer=None):
    """
    Exporta los rankings a un archivo Excel con hojas detalladas
    """
    if buffer is None:
        buffer = BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        for df, sheet_name in zip(rankings_dfs, sheet_names):
            # Crear una copia del DataFrame para no modificar el original
            detailed_df = df.copy()
            
            # Agregar columnas adicionales y formatear datos
            if 'puntaje_total' in detailed_df.columns:
                detailed_df['puntaje_porcentual'] = detailed_df['puntaje_total'].apply(
                    lambda x: f"{(x * 100):.2f}%"
                )
            
            # Reorganizar y renombrar columnas para mejor claridad
            column_mapping = {
                'nombre': 'Nombre Completo',
                'puntaje_total': 'Puntaje Total',
                'puntaje_porcentual': 'Porcentaje de Ajuste',
                'experiencia': 'Años de Experiencia',
                'educacion': 'Nivel Educativo',
                'habilidades_tecnicas': 'Habilidades Técnicas',
                'habilidades_blandas': 'Habilidades Blandas',
                'idiomas': 'Idiomas'
            }
            
            detailed_df = detailed_df.rename(columns=column_mapping)
            
            # Exportar a Excel con formato mejorado
            detailed_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Obtener el objeto worksheet
            worksheet = writer.sheets[sheet_name]
            
            # Ajustar el ancho de las columnas
            for idx, col in enumerate(detailed_df.columns):
                max_length = max(
                    detailed_df[col].astype(str).apply(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length
            
            # Agregar formato a los encabezados
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

    return buffer